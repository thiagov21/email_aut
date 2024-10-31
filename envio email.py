import os
import time
import win32com.client as win32
import openpyxl
import re

# Caminho da planilha de clientes
caminho_planilha = r"\\mao-fs01.technos.local\Arquivos\Departamental\PÓS VENDA\Laboratorio\COBRANÇA POSTOS Thiago V\automação\projeto email troca\DATA\email clientes.xlsx"

# Caminho do documento em PDF a ser anexado
caminho_anexo = r"\\mao-fs01.technos.local\Arquivos\Departamental\PÓS VENDA\Laboratorio\COBRANÇA POSTOS Thiago V\automação\projeto email troca\COMUNICADO\Manual_de_troca_site_Technos_Care 1.pdf"

link_info = f'https://technoscare.grupotechnos.com.br/Pergunta/DetalhesPergunta/41'

# Inicializa o Outlook
outlook = win32.Dispatch('outlook.application')

# Endereços de e-mail para cópia oculta (BCC)
emails_cco = "varfurado@gmail.com"

# Função para validar endereço de e-mail
def email_valido(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

# Tamanho do lote (quantidade de e-mails por envio)
lote_tamanho = 100

# Pausa entre lotes (em segundos)
pausa_lotes = 5

# Abrir a planilha
try:
    workbook = openpyxl.load_workbook(caminho_planilha)
    sheet = workbook.active

    # Total de linhas na planilha
    total_linhas = sheet.max_row

    # Loop para envio por lotes
    for inicio in range(1, total_linhas, lote_tamanho):
        fim = min(inicio + lote_tamanho, total_linhas)  # Define o fim do lote
        
        for row in sheet.iter_rows(min_row=inicio, max_row=fim, max_col=8, values_only=True):
            cliente_planilha = row[6]  # Coluna G (Cliente)
            email_destinatario = row[7]  # Coluna H (Email)
            relogio_cliente = row[3]  # Coluna D (Relógio)  
            OS_cliente = row[0]  # Coluna B (Ordem de Serviço)
            OS_troca = row[1]
            
            # Verificar se o e-mail é válido
            if email_valido(email_destinatario):
                # Criar um novo e-mail
                email = outlook.CreateItem(0)
                email.Subject = f"Grupo Technos - Ordem de Serviço"
                
                # Corpo do e-mail em HTML com link clicável
                corpo_email = f"""
                <p>Prezado(a) {cliente_planilha},</p>
                <p>Em referência ao serviço do relógio marca Technos, modelo {relogio_cliente},<br>
                informamos que, infelizmente, não será possível concluir o reparo do produto referente
                à Ordem de Serviço {OS_cliente} dentro do prazo legal de 30 (trinta) dias.<br>
                Gerando assim a OS de troca : {OS_troca}.</p>

                <p>Com o intuito de proporcionar a melhor experiência possível, gostaríamos de oferecer
                a substituição do produto por um modelo novo, conforme as opções abaixo:</p>

                <ul>
                    <li><strong>Garantia de venda:</strong> A substituição será feita sem qualquer custo adicional, com um novo relógio similar ao seu e uma nova garantia.</li>
                    <li><strong>Sem garantia:</strong> Haverá um custo para a substituição, porém, oferecemos um desconto referente ao valor do seu relógio no catálogo.</li>
                    <li><strong>Garantia parcial:</strong> Será cobrado apenas o custo da peça que não está coberta pela garantia.</li>
                </ul>

                <p>Para agilizar esse processo, disponibilizamos um catálogo no site <a href="https://technoscare.grupotechnos.com.br/">www.technoscare.com.br</a>,
                onde você poderá selecionar o modelo que mais lhe agradar. Os produtos disponíveis
                são equivalentes ao seu relógio original, tanto em preço quanto em características.</p>

                <p>É fundamental que essa escolha seja feita em até 05 (CINCO) dias úteis. Para auxiliá-lo(a),
                criamos um passo a passo que explica como realizar a seleção do novo produto.</p>

                <p>Para acessar o passo a passo <a href="{link_info}">CLIQUE AQUI</a>.</p>
                
                <p>Em anexo, também está o passo a passo para acessar o site.</p>

                <p>Após a confirmação de sua escolha, o novo produto será enviado com um prazo estimado
                de até 20 (VINTE) dias para entrega. Recomendamos que a escolha seja feita o quanto
                antes para que possamos atender você da maneira mais rápida possível.</p>

                <p>Caso nenhum dos modelos disponíveis atenda às suas expectativas, por favor, entre
                em contato conosco através do e-mail <a href="mailto:sac@grupotechnos.com.br">sac@grupotechnos.com.br</a>.
                Será possível solicitar uma nova seleção de modelos ou indicar até cinco opções das marcas do Grupo Technos
                que mais lhe agradem. Em caso de variação de preço, analisaremos as condições de
                disponibilidade e ajustaremos conforme necessário. Além disso, está disponível a
                opção de restituição do valor pago, conforme o artigo 18 do Código de Defesa do
                Consumidor (CDC).</p>

                <p>Lamentamos os eventuais transtornos e contamos com sua compreensão.</p>
                <p>Agradecemos sua preferência e confiança em nossos serviços.</p>
                
                <b>Este é um email automático. Por favor, não responda essa mensagem.</b>

                <p>Atenciosamente,<br>
                Grupo Technos<br>
                <a href="https://technoscare.grupotechnos.com.br/">www.technoscare.com.br</a></p>
                """


                email.HTMLBody = corpo_email  # Define o corpo do e-mail como HTML
                email.To = email_destinatario
                email.BCC = emails_cco 

                # Anexar o documento PDF
                email.Attachments.Add(caminho_anexo)

                # Enviar o e-mail
                email.Send()

                print(f"E-mail enviado para {cliente_planilha} ({email_destinatario}).")

            else:
                print(f"E-mail inválido para {cliente_planilha}: {email_destinatario}")

        # Pausa entre lotes
        print(f"Esperando {pausa_lotes} segundos antes de enviar o próximo lote...")
        time.sleep(pausa_lotes)

except Exception as e:
    print(f"Erro ao processar a planilha: {e}")

print(f"Todos os e-mails foram enviados. Total de: {total_linhas}")
